from openpyxl import load_workbook


# Function used to read the excel file and return first two columns
def readExcel(filepath):
    workbook = load_workbook(filepath)
    sheet = workbook.active

    # Variable "data" is a nested list used to store all rows of first two columns of excel file.
    # Variable "i" is used only to skip the first row of the excel file, assuming it contains the titles.
    data = []
    i = 0
    for row in sheet.iter_rows(values_only=True):
        if i != 0:
            data.append([row[0], row[1]])
        i += 1

    return data


# Function to check if the uploaded file is an excel file
def allowed_file(filename):
    # Checks if there is a dot in the file name, then checks if the string after the dot belongs
    # to excel file extensions.
    return "." in filename and filename.rsplit(".", 1)[1].lower() in {"xls", "xlsx"}
